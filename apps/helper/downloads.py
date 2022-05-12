from daisie.core.export_import import Downloader
import pandas as pd
import io
import uuid
import os

import os
import io
import uuid
import pandas as pd


from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font
from openpyxl.drawing.image import Image

class HelperDownload():

    @staticmethod
    def create_xlsx_for_download(data, fig=None,  kpi=None, keys = []):
        userid = uuid.uuid4()
        filepath = f'{userid}-excel.xlsx'
        
        if 'table' in keys:
            with pd.ExcelWriter(filepath) as writer:
                data.to_excel(writer, sheet_name='Daten')
        
        
        if 'kpi' in keys:
            if os.path.exists(filepath):
                workbook  = load_workbook(filename=filepath)
            else:
                workbook = Workbook()
                workbook.remove(workbook.active)
            worksheet = workbook.create_sheet(title='KPIs')

            # format creation
            format_kpi_header = Font(bold= True, size=14, name='Arial')
            worksheet.column_dimensions['B'].font = Font(bold= False, size=12, name='Arial')
            worksheet.column_dimensions['B'].width = 30
            worksheet.column_dimensions['C'].font = Font(bold= False, size=12, name='Arial')
            
            worksheet['A1'] = "KPIs"
            worksheet['A1'].font = format_kpi_header

            worksheet['B2'] = "Umsatz in EUR"
            worksheet['C2'] =  kpi.get('umsatz')
            
            
            worksheet['B4'] = "Durchschnittlicher Umsatz in EUR"
            worksheet['C4'] = kpi.get('durch_Umsatz')
            
            worksheet['B6'] = "Anzahl"
            worksheet['C6'] = kpi.get('anzahl')
            
            workbook.save(filepath)

        if 'plot' in keys:
            if os.path.exists(filepath):
                workbook  = load_workbook(filename=filepath)
            else:
                workbook = Workbook()
                workbook.remove(workbook.active)
            worksheet = workbook.create_sheet(title='Grafik')
            worksheet['A1'] = 'Grafik'

            with io.BytesIO() as fig_io:
                fig.write_image(file=fig_io, format='png', engine="kaleido")
                worksheet.add_image(Image(fig_io), 'B2')

                workbook.save(filepath)
        
        with open(filepath, 'rb') as xlsx_io: 
            mimetype = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            download = Downloader._create_downloader_href(xlsx_io, mimetype=mimetype)

        os.remove(filepath)
        return download 
    